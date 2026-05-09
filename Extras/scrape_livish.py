import re
import time
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

BASE_URL = "https://livish.com"
PAGES_TO_FETCH = 25

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

session = requests.Session()
session.headers.update(HEADERS)


def to_float(text):
    if not text:
        return None
    text = re.sub(r"[^\d.]", "", text.replace(",", ""))
    try:
        return float(text)
    except Exception:
        return None


def calc_discount(original, discounted):
    if original and discounted and original > 0:
        return (original - discounted) / original
    return None


def extract_product_cards(soup):
    products = []
    cards = soup.select(".t4s-product-wrapper")

    for card in cards:
        # URL
        a = card.select_one("a.t4s-full-width-link, a[data-pr-href]")
        href = a.get("href") if a else None
        product_url = urljoin(BASE_URL, href) if href else None
        if not product_url:
            continue

        # Image — lazy loaded via data-src on noscript img is reliable
        image_url = None
        noscript_img = card.select_one("noscript img")
        if noscript_img:
            src = noscript_img.get("src", "")
            if src.startswith("//"):
                src = "https:" + src
            elif src.startswith("/"):
                src = urljoin(BASE_URL, src)
            image_url = src
        else:
            img = card.select_one("img.t4s-product-main-img")
            if img:
                src = img.get("data-src") or img.get("src", "")
                if src and not src.startswith("data:"):
                    if src.startswith("//"):
                        src = "https:" + src
                    image_url = src

        # Name
        title_el = card.select_one(".t4s-product-title a, .t4s-product-title")
        name = title_el.get_text(" ", strip=True) if title_el else None

        # Prices — <del> = original, <ins> = discounted
        price_el = card.select_one(".t4s-product-price")
        original_price = None
        discounted_price = None
        if price_el:
            del_el = price_el.select_one("del")
            ins_el = price_el.select_one("ins")
            if del_el:
                original_price = to_float(del_el.get_text())
            if ins_el:
                discounted_price = to_float(ins_el.get_text())
            if not del_el and not ins_el:
                # Single price, no discount
                discounted_price = to_float(price_el.get_text())

        # Stock status from badge
        stock_status = "In Stock"
        badge = card.select_one(".t4s-product-badge")
        badge_text = badge.get_text(" ", strip=True).lower() if badge else ""
        if "sold out" in badge_text or "soldout" in badge_text:
            stock_status = "Out of Stock"

        products.append({
            "Image URL": image_url,
            "Product Name": name,
            "Original Price (KD)": original_price,
            "Discounted Price (KD)": discounted_price,
            "Discount %": calc_discount(original_price, discounted_price),
            "Stock Status": stock_status,
            "Product URL": product_url,
        })

    return products


def enrich_stock_status(product):
    if product["Stock Status"] != "Unknown":
        return product
    try:
        r = session.get(product["Product URL"], timeout=30)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        text = soup.get_text(" ", strip=True).lower()
        if "sold out" in text or "notify me when available" in text:
            product["Stock Status"] = "Out of Stock"
        elif "add to cart" in text:
            product["Stock Status"] = "In Stock"
    except Exception:
        pass
    return product


all_products = []
seen_urls = set()

for page_num in range(1, PAGES_TO_FETCH + 1):
    page_url = f"{BASE_URL}/collections/perfumes?page={page_num}&sort_by=best-selling"
    print(f"Fetching page {page_num}: {page_url}")
    try:
        r = session.get(page_url, timeout=30)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        page_products = extract_product_cards(soup)
        print(f"  → {len(page_products)} products on this page")
        for p in page_products:
            if p["Product URL"] not in seen_urls:
                p["Source Page"] = page_num
                all_products.append(p)
                seen_urls.add(p["Product URL"])
    except Exception as e:
        print(f"  ERROR: {e}")
    time.sleep(1.2)

print(f"\nFound {len(all_products)} unique products total")

# Enrich stock for any unknowns
unknowns = [p for p in all_products if p["Stock Status"] == "Unknown"]
if unknowns:
    print(f"Enriching stock for {len(unknowns)} unknown products...")
    for i, p in enumerate(all_products, start=1):
        all_products[i - 1] = enrich_stock_status(p)
        if i % 25 == 0:
            print(f"  Checked {i}/{len(all_products)}")
        time.sleep(0.4)

# Build Excel
wb = Workbook()
ws = wb.active
ws.title = "Livish Perfumes"

headers = [
    "Image URL",
    "Product Name",
    "Original Price (KD)",
    "Discounted Price (KD)",
    "Discount %",
    "Stock Status",
    "Product URL",
    "Source Page",
]
ws.append(headers)

for row in all_products:
    ws.append([
        row.get("Image URL"),
        row.get("Product Name"),
        row.get("Original Price (KD)"),
        row.get("Discounted Price (KD)"),
        row.get("Discount %"),
        row.get("Stock Status"),
        row.get("Product URL"),
        row.get("Source Page"),
    ])

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin_gray = Side(style="thin", color="D9D9D9")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(bottom=thin_gray)

for row in ws.iter_rows(min_row=2):
    row[2].number_format = '0.000 "KD"'
    row[3].number_format = '0.000 "KD"'
    row[4].number_format = "0.0%"

widths = {"A": 55, "B": 50, "C": 18, "D": 20, "E": 12, "F": 16, "G": 70, "H": 12}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

ws.freeze_panes = "A2"

output_file = "/Users/nishant/Desktop/onlyfats/livish_perfumes_first_25_pages.xlsx"
wb.save(output_file)
print(f"\nSaved: {output_file}")
