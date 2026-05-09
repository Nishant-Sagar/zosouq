"""
Livish Hair Care scraper
- Scrapes all pages of /collections/hair-care
- Saves product data to Excel
- Downloads all product images locally into subfolders by product name
"""

import os
import re
import time
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

BASE_URL = "https://livish.com"
COLLECTION_URL = "https://livish.com/collections/hair-care"
PAGES_TO_FETCH = 10
OUTPUT_EXCEL = Path("/Users/nishant/Desktop/onlyfats/livish_hair_care.xlsx")
OUTPUT_IMAGES = Path("/Users/nishant/Desktop/onlyfats/livish_hair_care_images")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

session = requests.Session()
session.headers.update(HEADERS)


def to_float(text):
    if not text:
        return None
    text = re.sub(r"[^\d.]", "", text.replace(",", ""))
    try:
        return float(text)
    except Exception:
        return None


def calc_discount(original, discounted):
    if original and discounted and original > 0:
        return (original - discounted) / original
    return None


def safe_folder_name(name: str) -> str:
    name = name.strip()
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name[:100].strip()


def image_ext(url: str) -> str:
    path = url.split("?")[0]
    ext = os.path.splitext(path)[1].lower()
    return ext if ext in (".jpg", ".jpeg", ".png", ".webp", ".gif") else ".jpg"


def scrape_collection_page(page_num: int):
    url = f"{COLLECTION_URL}?page={page_num}"
    r = session.get(url, timeout=30)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    products = []
    for card in soup.select(".t4s-product-wrapper"):
        a = card.select_one("a.t4s-full-width-link, a[data-pr-href]")
        href = a.get("href") if a else None
        if not href:
            continue
        product_url = urljoin(BASE_URL, href)

        # Image
        image_url = None
        noscript_img = card.select_one("noscript img")
        if noscript_img:
            src = noscript_img.get("src", "")
            if src.startswith("//"):
                src = "https:" + src
            elif src.startswith("/"):
                src = urljoin(BASE_URL, src)
            image_url = src
        else:
            img = card.select_one("img.t4s-product-main-img")
            if img:
                src = img.get("data-src") or img.get("src", "")
                if src and not src.startswith("data:"):
                    if src.startswith("//"):
                        src = "https:" + src
                    image_url = src

        # Name
        title_el = card.select_one(".t4s-product-title a, .t4s-product-title")
        name = title_el.get_text(" ", strip=True) if title_el else href.split("/")[-1]

        # Prices
        price_el = card.select_one(".t4s-product-price")
        original_price = discounted_price = None
        if price_el:
            del_el = price_el.select_one("del")
            ins_el = price_el.select_one("ins")
            if del_el:
                original_price = to_float(del_el.get_text())
            if ins_el:
                discounted_price = to_float(ins_el.get_text())
            if not del_el and not ins_el:
                discounted_price = to_float(price_el.get_text())

        # Stock
        stock_status = "In Stock"
        badge = card.select_one(".t4s-product-badge")
        if badge:
            if "sold out" in badge.get_text(" ", strip=True).lower():
                stock_status = "Out of Stock"

        products.append({
            "Image URL": image_url,
            "Product Name": name,
            "Original Price (KD)": original_price,
            "Discounted Price (KD)": discounted_price,
            "Discount %": calc_discount(original_price, discounted_price),
            "Stock Status": stock_status,
            "Product URL": product_url,
        })

    return products


def get_product_images(product_url: str):
    js_url = product_url.rstrip("/") + ".js"
    r = session.get(js_url, timeout=30)
    r.raise_for_status()
    data = r.json()
    images = data.get("images", [])
    result = []
    for img in images:
        if img.startswith("//"):
            img = "https:" + img
        img = re.sub(r'[?&]width=\d+', '', img)
        result.append(img)
    return result


def download_image(url: str, dest_path: Path):
    r = session.get(url, timeout=60, stream=True)
    r.raise_for_status()
    with open(dest_path, "wb") as f:
        for chunk in r.iter_content(chunk_size=8192):
            f.write(chunk)


# ── Phase 1: Scrape all pages ─────────────────────────────────────────────────

print("=" * 60)
print("PHASE 1: Scraping collection pages")
print("=" * 60)

all_products = []
seen_urls = set()

for page_num in range(1, PAGES_TO_FETCH + 1):
    print(f"Page {page_num}/{PAGES_TO_FETCH}...", end=" ", flush=True)
    try:
        products = scrape_collection_page(page_num)
        if not products:
            print("0 products — stopping (past last page)")
            break
        added = 0
        for p in products:
            if p["Product URL"] not in seen_urls:
                seen_urls.add(p["Product URL"])
                p["Source Page"] = page_num
                all_products.append(p)
                added += 1
        print(f"{added} new products (total: {len(all_products)})")
    except Exception as e:
        print(f"ERROR: {e}")
    time.sleep(1.0)

print(f"\nTotal unique products: {len(all_products)}")


# ── Phase 2: Save Excel ───────────────────────────────────────────────────────

print("\n" + "=" * 60)
print("PHASE 2: Saving Excel")
print("=" * 60)

wb = Workbook()
ws = wb.active
ws.title = "Livish Hair Care"

col_headers = [
    "Image URL", "Product Name", "Original Price (KD)",
    "Discounted Price (KD)", "Discount %", "Stock Status",
    "Product URL", "Source Page",
]
ws.append(col_headers)

for row in all_products:
    ws.append([
        row.get("Image URL"),
        row.get("Product Name"),
        row.get("Original Price (KD)"),
        row.get("Discounted Price (KD)"),
        row.get("Discount %"),
        row.get("Stock Status"),
        row.get("Product URL"),
        row.get("Source Page"),
    ])

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
thin_gray = Side(style="thin", color="D9D9D9")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(bottom=thin_gray)

for row in ws.iter_rows(min_row=2):
    row[2].number_format = '0.000 "KD"'
    row[3].number_format = '0.000 "KD"'
    row[4].number_format = "0.0%"

widths = {"A": 55, "B": 50, "C": 18, "D": 20, "E": 12, "F": 16, "G": 70, "H": 12}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

ws.freeze_panes = "A2"
wb.save(OUTPUT_EXCEL)
print(f"Saved: {OUTPUT_EXCEL}")


# ── Phase 3: Download images ──────────────────────────────────────────────────

print("\n" + "=" * 60)
print("PHASE 3: Downloading product images")
print("=" * 60)

OUTPUT_IMAGES.mkdir(parents=True, exist_ok=True)
total_images = 0
skipped = 0

for i, product in enumerate(all_products, start=1):
    name = product["Product Name"]
    product_url = product["Product URL"]
    folder = OUTPUT_IMAGES / safe_folder_name(name)

    print(f"[{i}/{len(all_products)}] {name}")

    try:
        images = get_product_images(product_url)
        if not images:
            print("  No images found")
            skipped += 1
            time.sleep(0.3)
            continue

        folder.mkdir(parents=True, exist_ok=True)

        for j, img_url in enumerate(images, start=1):
            ext = image_ext(img_url)
            dest = folder / f"image_{j:02d}{ext}"

            if dest.exists():
                print(f"  [{j}/{len(images)}] Already exists, skipping")
                continue

            try:
                download_image(img_url, dest)
                total_images += 1
                print(f"  [{j}/{len(images)}] Saved: {dest.name}")
            except Exception as e:
                print(f"  [{j}/{len(images)}] FAILED: {e}")

            time.sleep(0.15)

    except Exception as e:
        print(f"  ERROR: {e}")
        skipped += 1

    time.sleep(0.4)

print("\n" + "=" * 60)
print("DONE")
print(f"Products processed : {len(all_products) - skipped}")
print(f"Products skipped   : {skipped}")
print(f"Total images saved : {total_images}")
print(f"Excel              : {OUTPUT_EXCEL}")
print(f"Images folder      : {OUTPUT_IMAGES}")
print("=" * 60)
